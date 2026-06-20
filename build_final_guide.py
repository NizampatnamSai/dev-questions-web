import sys
sys.path.insert(0, '/Users/apple/Desktop/devquiz')

from study_data_html_css   import HTML_DATA, CSS_DATA
from study_data_js         import JS_DATA
from study_data_ts         import TS_DATA
from study_data_react_rn_next import REACT_DATA, RN_DATA, NEXT_DATA
from study_data_extra      import (RN_EXTRA, JS_EXTRA, REACT_EXTRA,
                                    NEXT_EXTRA, TS_EXTRA, HTML_EXTRA, CSS_EXTRA)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

CATS = {
    "HTML":          {"color": "C0392B", "icon": "🌐"},
    "CSS":           {"color": "1A5276", "icon": "🎨"},
    "JavaScript":    {"color": "7D6608", "icon": "⚡"},
    "TypeScript":    {"color": "1F618D", "icon": "🔷"},
    "React":         {"color": "0E6655", "icon": "⚛️"},
    "React Native":  {"color": "4A235A", "icon": "📱"},
    "Next.js":       {"color": "212F3D", "icon": "▲"},
}

DIFF_COLORS = {
    "Basic":        "C6EFCE",
    "Intermediate": "FFEB9C",
    "Advanced":     "FFC7CE",
    "Tricky":       "E1BEE7",
}

HEADER_BG = "1E293B"
thin = Side(style="thin", color="D1D5DB")
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(color):
    return PatternFill("solid", start_color=color, end_color=color)

ALL_DATA = {
    "HTML":         HTML_DATA  + HTML_EXTRA,
    "CSS":          CSS_DATA   + CSS_EXTRA,
    "JavaScript":   JS_DATA    + JS_EXTRA,
    "TypeScript":   TS_DATA    + TS_EXTRA,
    "React":        REACT_DATA + REACT_EXTRA,
    "React Native": RN_DATA    + RN_EXTRA,
    "Next.js":      NEXT_DATA  + NEXT_EXTRA,
}

INDEX_ROWS = []

for cat_name, rows in ALL_DATA.items():
    meta = CATS[cat_name]
    ws   = wb.create_sheet(title=cat_name)

    col_widths = [22, 22, 14, 44, 52, 64, 54]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    headers = ["Topic", "Subtopic", "Difficulty", "Key Concept",
               "Interview Question", "Answer / Explanation", "Code Snippet"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill      = fill(HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = bdr
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    for r, row in enumerate(rows, 2):
        topic, subtopic, diff, concept, question, answer, code = row
        values = [topic, subtopic, diff, concept, question, answer, code]
        for c, val in enumerate(values, 1):
            cell        = ws.cell(row=r, column=c, value=val)
            cell.border = bdr
            if c == 3:
                cell.fill      = fill(DIFF_COLORS.get(diff, "FFFFFF"))
                cell.font      = Font(name="Arial", size=10, bold=True, color="1E293B")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c == 7:
                cell.fill      = fill("0D1117")
                cell.font      = Font(name="Courier New", size=8, color="E2E8F0")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.font      = Font(name="Arial", size=10, color="1E293B")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if r % 2 == 0:
                    cell.fill  = fill("F8FAFC")
        ws.row_dimensions[r].height = 90

    INDEX_ROWS.append((cat_name, meta["icon"], meta["color"], len(rows)))

# ── Index / Cover ─────────────────────────────────────────────────────────────
cover = wb.create_sheet(title="📋 Index", index=0)
for col, w in zip("ABCDEFG", [5, 20, 9, 5, 46, 22, 22]):
    cover.column_dimensions[col].width = w

# Title
cover.merge_cells("A1:G1")
c = cover["A1"]
c.value      = "🚀  DevQuiz — Complete Interview Study Guide  (5-Year Experience Level)"
c.font       = Font(name="Arial", bold=True, size=18, color="FFFFFF")
c.fill       = fill(HEADER_BG)
c.alignment  = Alignment(horizontal="center", vertical="center")
cover.row_dimensions[1].height = 48

cover.merge_cells("A2:G2")
c = cover["A2"]
c.value     = "HTML  •  CSS  •  JavaScript  •  TypeScript  •  React  •  React Native  •  Next.js"
c.font      = Font(name="Arial", bold=True, size=13, color="6366F1")
c.alignment = Alignment(horizontal="center", vertical="center")
cover.row_dimensions[2].height = 28

total = sum(v for _, _, _, v in INDEX_ROWS)
cover.merge_cells("A3:G3")
c = cover["A3"]
c.value     = (f"✅  {total} topics covering every major area  ·  "
               "Basic · Intermediate · Advanced · Tricky Gotchas  ·  "
               "Designed for 5-year experience interviews")
c.font      = Font(name="Arial", size=11, color="475569")
c.alignment = Alignment(horizontal="center", vertical="center")
cover.row_dimensions[3].height = 22

# Table header
for col, label in zip("ABCDEFG", ["#","Technology","Topics","Icon","Coverage","Key additions","Jump"]):
    cell        = cover[f"{col}5"]
    cell.value  = label
    cell.font   = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    cell.fill   = fill(HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = bdr
cover.row_dimensions[5].height = 26

COVERAGE = {
    "HTML":         "Document structure, Semantics, Forms, Validation, Accessibility, ARIA, Canvas, SVG, Web Components, Shadow DOM, Security",
    "CSS":          "Box model, Specificity, Cascade layers, Flexbox, Grid, Subgrid, Animations, Transitions, Container queries, :has(), Nesting, View Transitions, oklch",
    "JavaScript":   "All loops, Arrays, Objects, Classes, Async/Await, Closures, Prototypes, Map/Set/WeakMap/Symbol, Regex, Design Patterns, Workers, WebSockets, ServiceWorkers, Security, Functional programming",
    "TypeScript":   "All type constructs, Generics, Utility types, Conditional/Mapped types, Decorators, Enums, Zod, Builder pattern, Exhaustive checks",
    "React":        "All hooks, Concurrent features, Error Boundaries, Testing (RTL+MSW), React Query, Jotai, React Hook Form, React 19 (use/useOptimistic), Fiber internals, Accessibility",
    "React Native": "Threads (JS/UI/Native), Push notifications (FCM/APNs), Gesture Handler, Reanimated, Testing (Jest+Detox), Native Modules, TurboModules, State mgmt, CodePush, EAS, Offline, Flipper",
    "Next.js":      "App Router, RSC, Server Actions, Caching, Middleware, Parallel/Intercepting routes, Auth.js, i18n, Error handling, Security headers, Bundle analysis, Vercel Edge",
}

KEY_ADD = {
    "HTML":         "dialog, srcset+sizes, loading=lazy, time element, contenteditable pitfalls",
    "CSS":          "View Transitions, CSS Anchor, oklch, color-mix(), aspect-ratio, @layer",
    "JavaScript":   "Design patterns, Web Workers, WebSockets, Service Workers, Functional programming, XSS/CORS security, Promise combinators",
    "TypeScript":   "Zod, Builder pattern, Type narrowing patterns, Path aliases, satisfies",
    "React":        "RTL testing, MSW mocking, React Query, Jotai, React Hook Form, use() hook, useOptimistic, Fiber, focus trap",
    "React Native": "Push notifications (full flow), JS/UI/Native threads, Gesture Handler, Detox E2E, Native Modules/TurboModules, CodePush, EAS Build, Flipper, Offline-first",
    "Next.js":      "Auth.js, Route protection, Next.js testing, Vercel Edge vs Serverless, i18n, error.tsx+global-error.tsx, CSP security headers, Bundle analyzer",
}

for i, (cat, icon, color, count) in enumerate(INDEX_ROWS, 1):
    r     = i + 5
    bg    = "F1F5F9" if i % 2 == 0 else "FFFFFF"
    for col_i, val in enumerate([i, cat, count, icon,
                                  COVERAGE.get(cat,""), KEY_ADD.get(cat,""), f"→ {cat}"], 1):
        cell        = cover.cell(row=r, column=col_i, value=val)
        cell.border = bdr
        cell.fill   = fill(bg)
        cell.font   = Font(name="Arial", size=10, color="1E293B")
        cell.alignment = Alignment(
            horizontal="center" if col_i in [1,3,4] else "left",
            vertical="center", wrap_text=True)
    # Bold tech name
    nc      = cover.cell(row=r, column=2)
    nc.font = Font(name="Arial", bold=True, size=11,
                   color=color if color not in ("212F3D",) else "334155")
    # Hyperlink
    lc       = cover.cell(row=r, column=7)
    lc.hyperlink = f"#{cat}!A1"
    lc.font  = Font(name="Arial", size=10, color="6366F1", underline="single")
    cover.row_dimensions[r].height = 36

# Legend
lr = len(INDEX_ROWS) + 8
cover.merge_cells(f"A{lr}:G{lr}")
lh = cover[f"A{lr}"]
lh.value     = "DIFFICULTY LEGEND"
lh.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
lh.fill      = fill(HEADER_BG)
lh.alignment = Alignment(horizontal="center")
cover.row_dimensions[lr].height = 24

LEGEND_DESC = {
    "Basic":        "Fundamental knowledge — expected from ALL candidates",
    "Intermediate": "Mid-level depth — core of most 3-5 year interviews",
    "Advanced":     "Deep expertise — differentiates senior/lead engineers",
    "Tricky":       "Gotchas, edge cases & misconceptions — what separates 8/10 from 10/10",
}
for i, (diff, dcolor) in enumerate(DIFF_COLORS.items()):
    r = lr + 1 + i
    c1 = cover.cell(row=r, column=1, value=diff)
    c1.fill = fill(dcolor)
    c1.font = Font(name="Arial", bold=True, size=10)
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.border = bdr
    cover.merge_cells(f"B{r}:G{r}")
    c2 = cover.cell(row=r, column=2, value=LEGEND_DESC[diff])
    c2.font   = Font(name="Arial", size=10)
    c2.border = bdr
    cover.row_dimensions[r].height = 20

output = "/Users/apple/Desktop/DevQuiz_Interview_Study_Guide.xlsx"
wb.save(output)
print(f"✅  Saved → {output}")
print(f"\n📊  Grand total: {total} topics\n")
for cat, icon, _, count in INDEX_ROWS:
    print(f"   {icon}  {cat:<18} {count} topics")
