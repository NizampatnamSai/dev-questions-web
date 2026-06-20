import sys
sys.path.insert(0, '/Users/apple/Desktop/devquiz')

from study_data_html_css import HTML_DATA, CSS_DATA
from study_data_js import JS_DATA
from study_data_ts import TS_DATA
from study_data_react_rn_next import REACT_DATA, RN_DATA, NEXT_DATA

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

CATS = {
    "HTML":          {"color": "C0392B", "light": "FDEDEC", "icon": "🌐"},
    "CSS":           {"color": "1A5276", "light": "D6EAF8", "icon": "🎨"},
    "JavaScript":    {"color": "7D6608", "light": "FEFDE7", "icon": "⚡"},
    "TypeScript":    {"color": "1F618D", "light": "D6EAF8", "icon": "🔷"},
    "React":         {"color": "0E6655", "light": "D5F5E3", "icon": "⚛️"},
    "React Native":  {"color": "4A235A", "light": "F5EEF8", "icon": "📱"},
    "Next.js":       {"color": "212F3D", "light": "EAECEE", "icon": "▲"},
}

DIFF_COLORS = {
    "Basic":        "C6EFCE",
    "Intermediate": "FFEB9C",
    "Advanced":     "FFC7CE",
    "Tricky":       "E1BEE7",
}

HEADER_BG = "1E293B"
thin = Side(style="thin", color="D1D5DB")
bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(color): return PatternFill("solid", start_color=color, end_color=color)

ALL_DATA = {
    "HTML": HTML_DATA,
    "CSS": CSS_DATA,
    "JavaScript": JS_DATA,
    "TypeScript": TS_DATA,
    "React": REACT_DATA,
    "React Native": RN_DATA,
    "Next.js": NEXT_DATA,
}

INDEX_ROWS = []

for cat_name, rows in ALL_DATA.items():
    meta = CATS[cat_name]
    ws = wb.create_sheet(title=cat_name)

    col_widths = [22, 22, 14, 42, 50, 62, 52]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    headers = ["Topic", "Subtopic", "Difficulty", "Key Concept",
               "Tricky Interview Question", "Answer / Explanation", "Code Snippet"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill = fill(HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    for r, row in enumerate(rows, 2):
        topic, subtopic, diff, concept, question, answer, code = row
        values = [topic, subtopic, diff, concept, question, answer, code]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = bdr
            if c == 3:  # Difficulty
                cell.fill = fill(DIFF_COLORS.get(diff, "FFFFFF"))
                cell.font = Font(name="Arial", size=10, bold=True, color="1E293B")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c == 7:  # Code
                cell.fill = fill("0D1117")
                cell.font = Font(name="Courier New", size=8, color="E2E8F0")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.font = Font(name="Arial", size=10, color="1E293B")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if r % 2 == 0:
                    cell.fill = fill("F8FAFC")
        ws.row_dimensions[r].height = 80

    INDEX_ROWS.append((cat_name, meta["icon"], meta["color"], len(rows)))

# ── Index sheet ─────────────────────────────────────────────────────────────
cover = wb.create_sheet(title="📋 Index", index=0)
for col, w in zip("ABCDEF", [6, 20, 10, 6, 44, 30]):
    cover.column_dimensions[col].width = w

cover.merge_cells("A1:F1")
c = cover["A1"]
c.value = "🚀  DevQuiz — Ultimate Interview Study Guide"
c.font = Font(name="Arial", bold=True, size=20, color="FFFFFF")
c.fill = fill(HEADER_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
cover.row_dimensions[1].height = 50

cover.merge_cells("A2:F2")
c = cover["A2"]
c.value = "HTML  •  CSS  •  JavaScript  •  TypeScript  •  React  •  React Native  •  Next.js"
c.font = Font(name="Arial", size=13, color="6366F1", bold=True)
c.alignment = Alignment(horizontal="center", vertical="center")
cover.row_dimensions[2].height = 30

# Stats row
cover.merge_cells("A3:F3")
total = sum(v for _, _, _, v in INDEX_ROWS)
c = cover["A3"]
c.value = f"Total topics: {total}   |   Covers: Basic · Intermediate · Advanced · Tricky Interview Gotchas"
c.font = Font(name="Arial", size=11, color="64748B")
c.alignment = Alignment(horizontal="center", vertical="center")
cover.row_dimensions[3].height = 24

# Header
for col, label in zip("ABCDEF", ["#", "Technology", "Topics", "Icon", "What's covered", "Jump to"]):
    cell = cover[f"{col}5"]
    cell.value = label
    cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    cell.fill = fill(HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = bdr
cover.row_dimensions[5].height = 28

coverage = {
    "HTML":         "Document structure, Semantics, Forms, Accessibility, Canvas, SVG, Web Components, Shadow DOM",
    "CSS":          "Box model, Specificity, Flexbox, Grid, Subgrid, Animations, Container queries, CSS nesting, :has()",
    "JavaScript":   "Loops, Arrays, Objects, Classes, Promises, Async/Await, Closures, Prototypes, Map/Set, Regex, DOM",
    "TypeScript":   "Types, Generics, Utility types, Conditional types, Mapped types, Decorators, Enums, satisfies",
    "React":        "All hooks, memo, Suspense, Concurrent, Error Boundaries, Portals, HOC, Compound components",
    "React Native": "Core components, Navigation, Animations, Reanimated, New Architecture, Permissions, Storage",
    "Next.js":      "App Router, RSC, Server Actions, Caching, Middleware, Parallel/Intercepting routes, Metadata",
}

for i, (cat, icon, color, count) in enumerate(INDEX_ROWS, 1):
    r = i + 5
    cells_vals = [i, cat, count, icon, coverage.get(cat, ""), f"→ {cat}"]
    for c_idx, val in enumerate(cells_vals, 1):
        cell = cover.cell(row=r, column=c_idx, value=val)
        cell.border = bdr
        cell.alignment = Alignment(horizontal="center" if c_idx in [1,3,4] else "left", vertical="center")
        if i % 2 == 0:
            cell.fill = fill("F1F5F9")
    # Highlight technology name
    name_cell = cover.cell(row=r, column=2)
    name_cell.font = Font(name="Arial", bold=True, size=11,
                          color=color if color not in ("000000", "212F3D") else "334155")
    # Hyperlink
    link_cell = cover.cell(row=r, column=6)
    link_cell.hyperlink = f"#{cat}!A1"
    link_cell.font = Font(name="Arial", size=10, color="6366F1", underline="single")
    cover.row_dimensions[r].height = 26

# Legend
lr = len(INDEX_ROWS) + 8
cover.merge_cells(f"A{lr}:F{lr}")
lh = cover[f"A{lr}"]
lh.value = "DIFFICULTY LEGEND"
lh.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
lh.fill = fill(HEADER_BG)
lh.alignment = Alignment(horizontal="center")
cover.row_dimensions[lr].height = 26

descs = {
    "Basic":        "Fundamental knowledge — expected from all candidates",
    "Intermediate": "Solid mid-level understanding — most interview questions",
    "Advanced":     "Deep expertise — senior / lead developer questions",
    "Tricky":       "Gotchas, edge cases, common misconceptions — separates strong candidates",
}
for i, (diff, dcolor) in enumerate(DIFF_COLORS.items()):
    r = lr + 1 + i
    c1 = cover.cell(row=r, column=1, value=diff)
    c1.fill = fill(dcolor)
    c1.font = Font(name="Arial", bold=True, size=10)
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.border = bdr
    cover.merge_cells(f"B{r}:F{r}")
    c2 = cover.cell(row=r, column=2, value=descs[diff])
    c2.font = Font(name="Arial", size=10)
    c2.border = bdr
    cover.row_dimensions[r].height = 22

output = "/Users/apple/Desktop/DevQuiz_Interview_Study_Guide.xlsx"
wb.save(output)
print(f"✅  Saved: {output}")
print(f"📊  Total topics: {total}")
for cat, icon, _, count in INDEX_ROWS:
    print(f"   {icon} {cat}: {count} topics")
